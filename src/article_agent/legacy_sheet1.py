from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .paths import LABELS_DIR
from .schemas import ParsedDocument, StudyRecord

SHEET1_COLUMN_COUNT = 91


def _text(doc: ParsedDocument) -> str:
    return " ".join(c.text for c in doc.chunks)


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _match(pattern: str, text: str, flags=re.I) -> str | None:
    m = re.search(pattern, text, flags)
    return _norm(m.group(1)) if m else None


def _contains(text: str, *needles: str) -> bool:
    low = text.lower()
    return any(n.lower() in low for n in needles)


def _country_category(country: str) -> str:
    eastern = {"china", "south korea", "korea", "japan", "taiwan", "hong kong"}
    western = {"spain", "australia", "united states", "usa", "germany", "turkey", "brazil", "iran"}
    low = country.lower()
    if low in eastern:
        return "1"
    if low in western:
        return "2"
    return "3" if country not in {"", "NR"} else "NR"


def _disease_system(disease: str) -> str:
    low = disease.lower()
    if any(k in low for k in ["fibromyalgia", "pain", "fatigue", "musculoskeletal"]):
        return "1"
    if any(k in low for k in ["stroke", "dysphagia"]):
        return "2"
    if any(k in low for k in ["rhinitis", "allergic"]):
        return "8"
    if any(k in low for k in ["urinary", "abortion"]):
        return "4"
    return "NR"


def _control_type(intervention: str, control: str) -> str:
    c = control.lower()
    i = intervention.lower()
    if "sham" in c and "usual" in c:
        return "5"
    if "sham" in c or "placebo" in c:
        return "3"
    if "usual" in c or "standard" in c:
        return "5"
    if "wait" in c:
        return "1"
    return "NR"


def _comparison_type(intervention: str, control: str) -> str:
    i = intervention.lower()
    c = control.lower()
    if "usual" in i and ("sham" in c or "placebo" in c) and "usual" in c:
        return "4"
    if "usual" in i and "usual" in c:
        return "3"
    if "sham" in c or "placebo" in c:
        return "2"
    if "wait" in c or "no treatment" in c:
        return "1"
    return "NR"


def _random_text(text: str) -> str:
    patterns = [
        r"([^.]*(?:random allocation sequence|computer-generated|random number table|centralised telephone|centralized telephone|randomisation service|randomization service)[^.]*\.)",
        r"([^.]*(?:randomly assigned|randomised|randomized)[^.]*\.)",
    ]
    for pat in patterns:
        value = _match(pat, text)
        if value:
            return value
    return "NR"


def _random_code(random_text: str) -> str:
    low = random_text.lower()
    if "random number" in low:
        return "1"
    if "computer" in low or "sas" in low or "random allocation sequence" in low:
        return "2"
    if random_text != "NR":
        return "9"
    return "8"


def _allocation_text(text: str) -> str:
    patterns = [
        r"([^.]*(?:allocation was concealed|opaque sealed envelopes|sealed opaque envelopes|centralised telephone|centralized telephone|independent clinical trials unit|allocation process)[^.]*\.)",
        r"([^.]*(?:concealed)[^.]*\.)",
    ]
    for pat in patterns:
        value = _match(pat, text)
        if value:
            return value
    return "NR"


def _allocation_code(value: str) -> str:
    low = value.lower()
    if "central" in low or "independent clinical trials unit" in low:
        return "1"
    if "opaque" in low and "sealed" in low:
        return "2"
    if "sealed" in low:
        return "3"
    if value != "NR":
        return "6"
    return "5"


def _yes_no_unclear(text: str, yes_patterns: list[str]) -> str:
    low = text.lower()
    if any(p in low for p in yes_patterns):
        return "1"
    return "3"


def _longest_followup(text: str) -> tuple[str, str, str]:
    candidates: list[tuple[int, str, str]] = []
    for value, unit in re.findall(r"(\d+)\s*(weeks?|months?|years?)", text, flags=re.I):
        v = int(value)
        u = unit.lower()
        days = v * (365 if u.startswith("year") else 30 if u.startswith("month") else 7)
        candidates.append((days, value, u))
    if not candidates:
        return "NR", "NR", "NR"
    _, value, unit = max(candidates, key=lambda x: x[0])
    unit_code = "4" if unit.startswith("week") else "2" if unit.startswith("month") else "3" if unit.startswith("year") else "NR"
    return f"{value} {unit}", value, unit_code


def _sex_values(text: str) -> tuple[str, str]:
    m = re.search(r"(\d+)\s*/\s*(\d+)", text)
    return (m.group(1), m.group(2)) if m else ("NR", "NR")


def _sample_values(study: StudyRecord, text: str) -> tuple[str, str, str, str, str]:
    total = study.randomized_n.value if study.randomized_n.value not in (None, "NR", "") else "NR"
    groups = re.search(r"n\s*=\s*(\d{2,4})\s+n\s*=\s*(\d{2,4})", text, flags=re.I)
    if groups:
        return str(total), groups.group(1), groups.group(2), groups.group(1), groups.group(2)
    return str(total), "NR", "NR", "NR", "NR"


def build_legacy_sheet1_record(doc: ParsedDocument, study: StudyRecord) -> dict[str, Any]:
    text = _text(doc)
    intervention = str(study.intervention_name.value or "NR")
    control = str(study.control_name.value or "NR")
    disease = str(study.disease_name.value or "NR")
    country = str(study.country.value or "NR")
    random_text = _random_text(text)
    alloc_text = _allocation_text(text)
    follow_text, follow_value, follow_unit = _longest_followup(text)
    randomized_total, rand_i, rand_c, analyzed_i, analyzed_c = _sample_values(study, text)

    row = {i: "NR" for i in range(1, SHEET1_COLUMN_COUNT + 1)}
    row.update(
        {
            1: study.study_id,
            3: study.extraction_status,
            4: "2" if "pain" in str(study.primary_outcome.value).lower() else "0",
            6: "1",
            7: study.title.value,
            8: study.year.value,
            9: study.language.value,
            10: study.journal.value,
            13: _match(r"trial\s+([A-Z][A-Za-z\-]+\s+[A-Z][A-Za-z\-]+)", text) or "NR",
            17: "1" if _contains(text, "trial registration", "ClinicalTrials.gov", "ISRCTN", "ANZCTR") else "2",
            18: disease,
            19: _disease_system(disease),
            20: "2" if _contains(disease, "chronic", "fibromyalgia", "fatigue") else "NR",
            21: country,
            22: _country_category(country),
            23: _match(r"Conducted in (\w+) primary care centres", text) or "NR",
            24: intervention,
            25: control,
            26: _control_type(intervention, control),
            28: _comparison_type(intervention, control),
            29: random_text,
            30: _random_code(random_text),
            31: alloc_text,
            32: _allocation_code(alloc_text),
            33: "1" if _contains(text, "stratified") else "2" if _contains(text, "block random") else "0",
            34: _yes_no_unclear(text, ["participants were blinded", "blinded to participants", "participant-blinded"]),
            35: _yes_no_unclear(text, ["practitioners were blinded", "personnel were blinded"]),
            36: "1" if _contains(text, "data analysts") else "3",
            40: "1" if _contains(intervention, "acupuncture") else "9",
            41: "1",
            42: "3" if _contains(intervention, "individual") else "4",
            45: _match(r"(one session per week|\d+\s*(?:to|-)?\s*\d*\s*times a week|\d+\s*times per week)", text) or "NR",
            46: _match(r"(\d+)\s*(?:times per week|session per week)", text) or ("1" if _contains(text, "one session per week") else "NR"),
            47: "2" if _contains(text, "week") else "NR",
            49: _match(r"(\d+)\s+weeks", text) or "NR",
            50: _match(r"(\d+)\s+weeks", text) or "NR",
            51: "2" if _contains(text, "week") else "NR",
            52: study.treatment_sessions.value,
            56: _match(r"depth[^.]*?between\s+([\d\-– ]+\s*mm)", text) or "NR",
            59: _match(r"(\d+)\s*min", text) or "NR",
            60: _match(r"(\d+)\s*min", text) or "NR",
            61: "1" if _contains(text, "min") else "NR",
            68: follow_text,
            69: follow_value,
            70: follow_unit,
            80: randomized_total,
            81: rand_i,
            82: rand_c,
            84: analyzed_i,
            85: analyzed_c,
            88: "1" if _contains(text, "missing", "dropout", "withdraw") else "3",
            89: _yes_no_unclear(text, ["intention-to-treat", "modified intention-to-treat"]),
            90: "4" if _contains(text, "last observation carried forward", "locf") else "10" if _contains(text, "mixed model", "mmrm") else "13",
        }
    )
    return {str(k): v for k, v in row.items()}


def find_legacy_sheet1_template(year: str | None) -> Path | None:
    if year:
        matches = sorted(LABELS_DIR.glob(f"{year}*.xlsx"))
        if matches:
            return matches[0]
    matches = sorted(LABELS_DIR.glob("*.xlsx"))
    return matches[0] if matches else None


def _find_sheet1(wb):
    candidates = [ws for ws in wb.worksheets if ws.max_column >= 80]
    return candidates[0] if candidates else wb.worksheets[0]


def write_legacy_sheet1(template_path: Path | None, output_path: Path, records: list[dict[str, Any]]) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Sheet1_auto"

    if template_path and template_path.exists():
        src_wb = load_workbook(template_path, read_only=False, data_only=False)
        src_ws = _find_sheet1(src_wb)
        max_col = max(SHEET1_COLUMN_COUNT, src_ws.max_column)
        for row_idx in [1, 2]:
            for col in range(1, max_col + 1):
                out_ws.cell(row=row_idx, column=col).value = src_ws.cell(row=row_idx, column=col).value
                out_ws.column_dimensions[get_column_letter(col)].width = src_ws.column_dimensions[get_column_letter(col)].width or 12
    else:
        for col in range(1, SHEET1_COLUMN_COUNT + 1):
            out_ws.cell(row=2, column=col).value = f"col_{col}"

    for row_offset, record in enumerate(records, start=3):
        for col in range(1, SHEET1_COLUMN_COUNT + 1):
            out_ws.cell(row=row_offset, column=col).value = record.get(str(col), "NR")

    out_wb.save(output_path)
    return output_path


def _row_dict(ws, row_index: int) -> dict[str, Any]:
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    return {headers[i]: ws.cell(row=row_index, column=i + 1).value for i in range(len(headers)) if headers[i]}


def _first_data_row(wb, sheet: str) -> dict[str, Any]:
    if sheet not in wb.sheetnames or wb[sheet].max_row < 2:
        return {}
    return _row_dict(wb[sheet], 2)


def _code_from_template(value: Any, mapping: dict[str, str], default: str = "NR") -> str:
    return mapping.get(str(value or "").lower(), default)


def build_legacy_sheet1_records_from_optimized(optimized_workbook: Path) -> list[dict[str, Any]]:
    wb = load_workbook(optimized_workbook, data_only=True)
    records: list[dict[str, Any]] = []
    study_rows = wb["01_Study"] if "01_Study" in wb.sheetnames else None
    if not study_rows:
        return records
    for row_index in range(2, study_rows.max_row + 1):
        study = _row_dict(study_rows, row_index)
        if not study.get("study_id"):
            continue
        arms = [_row_dict(wb["02_Arms"], r) for r in range(2, wb["02_Arms"].max_row + 1)] if "02_Arms" in wb.sheetnames else []
        comparisons = [_row_dict(wb["03_Comparisons"], r) for r in range(2, wb["03_Comparisons"].max_row + 1)] if "03_Comparisons" in wb.sheetnames else []
        method = _first_data_row(wb, "04_Methods")
        acu = _first_data_row(wb, "05_Acupuncture")
        outcome = _first_data_row(wb, "06_Outcomes")
        intervention_arm = next((a for a in arms if a.get("arm_role") == "intervention"), arms[0] if arms else {})
        control_arm = next((a for a in arms if a.get("arm_role") in {"control", "sham", "usual_care"}), arms[1] if len(arms) > 1 else {})
        comparison = comparisons[0] if comparisons else {}
        row = {i: "NR" for i in range(1, SHEET1_COLUMN_COUNT + 1)}
        row.update({
            1: study.get("study_id"),
            3: study.get("extraction_status") or "completed_with_review",
            4: "2" if outcome.get("patient_important_category") == "pain" else "0",
            6: "1" if study.get("eligibility_status") in {None, "", "include"} else study.get("eligibility_status"),
            7: study.get("title") or "NR",
            8: study.get("year") or "NR",
            9: study.get("language") or "NR",
            10: study.get("journal") or "NR",
            13: study.get("first_author") or "NR",
            14: study.get("corresponding_author_email") or "NR",
            17: "1" if method.get("protocol_registration") == "reported" else "2",
            18: study.get("disease_name") or "NR",
            19: _disease_system(str(study.get("disease_name") or "")),
            20: "2" if "chronic" in str(study.get("disease_name") or "").lower() or "fibromyalgia" in str(study.get("disease_name") or "").lower() else "NR",
            21: study.get("country") or "NR",
            22: _country_category(str(study.get("country") or "NR")),
            23: study.get("center_count") or "NR",
            24: intervention_arm.get("intervention_components") or intervention_arm.get("arm_name") or "NR",
            25: control_arm.get("intervention_components") or control_arm.get("arm_name") or "NR",
            26: _control_type(str(intervention_arm.get("intervention_components") or ""), str(control_arm.get("intervention_components") or "")),
            28: _code_from_template(comparison.get("comparison_type_code"), {"A": "2", "B": "4", "C": "3", "D": "1", "E": "NR"}, _comparison_type(str(intervention_arm.get("intervention_components") or ""), str(control_arm.get("intervention_components") or ""))),
            29: method.get("random_sequence_text") or "NR",
            30: _code_from_template(method.get("random_sequence_code"), {"random_number_table": "1", "computer": "2", "central_service": "9", "nr": "8"}, "8"),
            31: method.get("allocation_concealment_text") or "NR",
            32: _code_from_template(method.get("allocation_concealment_code"), {"central_service": "1", "sealed_opaque_envelope": "2", "independent_unit": "1", "nr": "5"}, "5"),
            34: "1" if method.get("participant_blinding_code") == "yes" else "3",
            35: "1" if method.get("personnel_blinding_code") == "yes" else "3",
            36: "1" if method.get("statistician_blinding_code") == "yes" else "3",
            40: _code_from_template(acu.get("acupuncture_modality"), {"manual": "1", "electroacupuncture": "2", "auricular": "6", "saam": "9"}, "9"),
            41: "1" if acu.get("stimulation_type") in {"manual", "NR", None, ""} else "2",
            42: _code_from_template(acu.get("point_selection_scheme"), {"fixed": "1", "semi_standardized": "2", "individualized": "3", "nr": "4"}, "4"),
            45: acu.get("frequency_per_week") or "NR",
            46: acu.get("frequency_per_week") or "NR",
            47: "2" if acu.get("frequency_per_week") not in {None, "", "NR"} else "NR",
            49: acu.get("treatment_duration_weeks") or "NR",
            50: acu.get("treatment_duration_weeks") or "NR",
            51: "2" if acu.get("treatment_duration_weeks") not in {None, "", "NR"} else "NR",
            52: acu.get("total_sessions") or "NR",
            56: acu.get("insertion_depth") or "NR",
            59: acu.get("retention_time_min") or "NR",
            60: acu.get("retention_time_min") or "NR",
            61: "1" if acu.get("retention_time_min") not in {None, "", "NR"} else "NR",
            80: intervention_arm.get("sample_randomized") or "NR",
            81: intervention_arm.get("sample_randomized") or "NR",
            82: control_arm.get("sample_randomized") or "NR",
            88: "1" if method.get("missing_data_method") not in {None, "", "NR"} else "3",
            89: "1" if method.get("primary_analysis_set") == "intention_to_treat" else "4",
            90: _code_from_template(method.get("missing_data_method"), {"locf": "4", "mixed_model": "10", "nr": "13"}, "13"),
        })
        records.append({str(k): v for k, v in row.items()})
    return records
