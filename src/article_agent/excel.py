from __future__ import annotations

from pathlib import Path
from shutil import copyfile
from typing import Iterable

from openpyxl import load_workbook

from .evaluation import EVALUATION_HEADERS
from .schemas import ArmRecord, ComparisonRecord, EvidenceSpan, OutcomeRecord, ReviewRecord, StudyRecord


def _header_map(ws) -> dict[str, int]:
    return {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value not in (None, "")}


def _write_row(ws, row_index: int, data: dict) -> None:
    headers = _header_map(ws)
    for key, value in data.items():
        if key in headers:
            ws.cell(row=row_index, column=headers[key]).value = value


def _clear_data_rows(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def _ensure_sheet(wb, sheet_name: str, headers: list[str]):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    _clear_data_rows(ws)
    if ws.max_row == 0:
        ws.append(headers)
    else:
        existing = [cell.value for cell in ws[1]]
        if existing[:len(headers)] != headers:
            for idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=idx).value = header
    return ws


def write_workbook(template_path: Path, output_path: Path, studies: list[StudyRecord], arms: list[ArmRecord], comparisons: list[ComparisonRecord], outcomes: list[OutcomeRecord], evidence: list[EvidenceSpan], review: list[ReviewRecord], methods: list[dict] | None = None, acupuncture: list[dict] | None = None, evaluation_rows: list[dict] | None = None) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    copyfile(template_path, output_path)
    wb = load_workbook(output_path)
    for sheet in ["01_Study", "02_Arms", "03_Comparisons", "04_Methods", "05_Acupuncture", "06_Outcomes", "08_Evidence", "09_Needs_Review"]:
        if sheet in wb.sheetnames:
            _clear_data_rows(wb[sheet])

    ws = wb["01_Study"]
    for i, s in enumerate(studies, start=2):
        _write_row(ws, i, {
            "study_id": s.study_id,
            "source_pdf": s.source_pdf,
            "extraction_status": s.extraction_status,
            "title": s.title.value,
            "year": s.year.value,
            "journal": s.journal.value,
            "language": s.language.value,
            "first_author": s.first_author.value,
            "corresponding_author": s.corresponding_author.value,
            "corresponding_author_email": s.corresponding_author_email.value,
            "country": s.country.value,
            "country_category": s.country_category.value,
            "setting": s.setting.value,
            "disease_name": s.disease_name.value,
            "disease_system": s.disease_system.value,
            "acute_or_chronic": s.acute_or_chronic.value,
            "surgical_or_procedural": s.surgical_or_procedural.value,
            "study_design": s.study_design.value,
            "center_count": s.center_count.value,
            "recruitment_start": s.recruitment_start.value,
            "recruitment_end": s.recruitment_end.value,
            "registration_id": s.doi.value,
            "funding": s.funding.value,
            "conflict_of_interest": s.conflict_of_interest.value,
            "eligibility_status": s.eligibility_status.value,
            "exclusion_reason": s.exclusion_reason.value,
            "notes": s.notes,
        })

    ws = wb["02_Arms"]
    for i, a in enumerate(arms, start=2):
        _write_row(ws, i, a.model_dump())

    ws = wb["03_Comparisons"]
    for i, c in enumerate(comparisons, start=2):
        _write_row(ws, i, c.model_dump())


    if methods:
        ws = wb["04_Methods"]
        for i, row in enumerate(methods, start=2):
            _write_row(ws, i, row)

    if acupuncture:
        ws = wb["05_Acupuncture"]
        for i, row in enumerate(acupuncture, start=2):
            _write_row(ws, i, row)

    ws = wb["06_Outcomes"]
    for i, o in enumerate(outcomes, start=2):
        _write_row(ws, i, o.model_dump())

    ws = wb["08_Evidence"]
    for i, e in enumerate(evidence, start=2):
        _write_row(ws, i, e.model_dump())

    ws = wb["09_Needs_Review"]
    for i, r in enumerate(review, start=2):
        _write_row(ws, i, r.model_dump())

    if evaluation_rows is not None:
        ws = _ensure_sheet(wb, "Evaluation", EVALUATION_HEADERS)
        for i, row in enumerate(evaluation_rows, start=2):
            _write_row(ws, i, row)
    wb.save(output_path)
    return output_path


