from __future__ import annotations

import json
import hashlib
import os
import re
import sys
import time
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(ROOT / "src"))

from article_agent.models import OpenAICompatibleClient


SHEET1_FIELDS = {
    "title", "publication_year", "journal", "first_author", "country",
    "intervention", "control", "control_type_transformed", "random_sequence_method",
    "random_sequence_class", "allocation_concealment", "allocation_concealment_class",
    "participant_blinding", "outcome_assessor_blinding", "treatment_frequency_raw",
    "treatment_frequency_value", "treatment_frequency_unit", "treatment_duration_raw",
    "treatment_duration_value", "treatment_duration_unit", "total_sessions", "deqi",
    "randomized_sample_intervention_raw", "randomized_sample_control_raw", "total_randomized",
    "primary_analysis", "missing_data_method",
}


def _article_rows(sheet, article_id: str, id_column: int) -> list[int]:
    """Return every gold row belonging to an article, including multi-arm suffixes."""
    rows = []
    for row in range(1, sheet.max_row + 1):
        value = str(sheet.cell(row, id_column).value or "").strip()
        if value == article_id or value.startswith(f"{article_id}-"):
            rows.append(row)
    return rows


def sheet1_gold(article_id: str) -> dict:
    mapping = json.loads((ROOT / "registry/legacy-excel/sheet1-mapping.json").read_text(encoding="utf-8"))
    workbook = load_workbook(ROOT / "Datas/label/2015-6篇-0813.xlsx", data_only=True)
    sheet = workbook["Sheet1"]
    rows = _article_rows(sheet, article_id, 1)
    if not rows:
        return {}
    result = {}
    for item in mapping["columns"]:
        field = item.get("canonicalFieldId")
        if field in SHEET1_FIELDS:
            values = []
            for row in rows:
                value = sheet.cell(row, item["columnIndex"]).value
                if value not in values:
                    values.append(value)
            result[field] = {
                "value": values[0] if len(values) == 1 else values,
                "excel_column": item["column"],
                "codebook": item.get("legacyCode"),
            }
    return result


def sheet3_gold(article_id: str) -> list[dict]:
    workbook = load_workbook(ROOT / "Datas/label/2015-6篇-0813.xlsx", data_only=True)
    sheet = workbook["Sheet3"]
    records = []
    for row in _article_rows(sheet, article_id, 1):
        records.append({
            str(sheet.cell(5, column).value or f"column_{column}"): sheet.cell(row, column).value
            for column in range(1, sheet.max_column + 1)
            if sheet.cell(row, column).value not in (None, "")
        })
    return records


def evidence_window(text: str, needle: str, radius: int = 2200) -> str:
    # Keep the helper lossless for callers that still use it directly.  The
    # radius remains a compatibility argument, not a correctness boundary.
    del radius
    index = text.find(needle)
    if index < 0:
        return text
    return text


def evidence_windows(text: str, needles: list[str], radius: int = 1000, limit: int = 7000) -> str:
    """Return complete routed evidence; never silently clip a section.

    The parameters are retained for compatibility with older report builders.
    Correctness-critical evaluation now partitions by table/row when needed,
    so a keyword window must not hide an outcome or a missing-data statement.
    """

    del needles, radius, limit
    return text


def postprocessed_outcome_candidate(extraction_dir: Path, raw_outcomes: dict) -> tuple[dict, bool]:
    """Build a compact candidate from post-extraction outcome annotations.

    The source outcome values remain nested under ``source_outcome`` and are
    never replaced.  Only the LLM's normalized name/instrument/timepoint and
    comparison annotations are added for the optional postprocessed score.
    Gold rows and gold row IDs are deliberately not copied into the candidate
    payload, so the evaluator still judges against the independent gold input.
    """
    path = extraction_dir / "outcomes.postprocessed.json"
    if not path.exists():
        return raw_outcomes, False
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        records = payload.get("records")
    except (OSError, TypeError, ValueError):
        return raw_outcomes, False
    # A partial/stale annotation file must never shrink the correctness
    # candidate.  The raw extraction is authoritative whenever the
    # post-processing stage did not cover every current source record.
    raw_records = raw_outcomes.get("outcomes", []) if isinstance(raw_outcomes, dict) else []
    expected_count = len(raw_records) if isinstance(raw_records, list) else 0
    if (
        payload.get("status") != "success"
        or not isinstance(records, list)
        or not records
        or (expected_count and len(records) != expected_count)
    ):
        return raw_outcomes, False
    compact_records = []
    for record in records:
        if not isinstance(record, dict):
            continue
        source = record.get("source_outcome")
        if not isinstance(source, dict):
            continue
        source = dict(source)
        compact_records.append({
            "source_index": record.get("source_index"),
            "source_outcome": source,
            "normalized_outcome_name": record.get("normalized_outcome_name", "NR"),
            "normalized_measurement_instrument": record.get("normalized_measurement_instrument", "NR"),
            "normalized_timepoint": record.get("normalized_timepoint", "NR"),
            "comparison_relation": record.get("comparison_relation", "NR"),
            "duplicate_group": record.get("duplicate_group"),
            "conflict_group_id": record.get("conflict_group_id"),
            "postprocess_conflict_status": record.get("conflict_status", "unresolved"),
            "postprocess_conflict_fields": record.get("conflict_fields", []),
            "annotation_status": record.get("annotation_status", record.get("conflict_status", "unresolved")),
        })
    if not compact_records:
        return raw_outcomes, False
    return {"outcomes": compact_records}, True


def compact_outcome_evaluation_candidate(candidate: dict, gold_rows: list[dict]) -> dict:
    """Return every outcome record for correctness-critical evaluation.

    The legacy function name remains to avoid breaking report code, but this
    projection is now lossless and source-first.  Per-row evaluation below
    partitions the records at the request boundary rather than dropping rows.
    """

    records = candidate.get("outcomes", []) if isinstance(candidate, dict) else []
    records = [record for record in records if isinstance(record, dict)] if isinstance(records, list) else []
    table_counts: dict[str, int] = {}
    for item in records:
        source = item.get("source_outcome") if isinstance(item.get("source_outcome"), dict) else item
        table_id = str(source.get("table_id") or "NR")
        table_counts[table_id] = table_counts.get(table_id, 0) + 1
    return {
        "record_count": len(records),
        "table_counts": table_counts,
        "outcome_records": records,
        "gold_reference_rows": [dict(row) for row in gold_rows if isinstance(row, dict)],
        "lossless": True,
        "note": "每条 source_outcome、证据、arm、比较和冲突注释均保留；评估按 source_index 分片。",
    }


def compact_gold_rows_for_evaluation(rows: list[dict], max_rows: int = 8, max_chars: int = 2800) -> list[dict]:
    """Return every Gold row/column; size arguments are compatibility-only."""

    del max_rows, max_chars
    return [dict(row) for row in rows if isinstance(row, dict)]


def call_judge(
    client: OpenAICompatibleClient,
    payload: dict,
    output: Path,
    retries: int = 4,
    request_delay_seconds: float = 0.01,
) -> dict:
    disable_cache = os.getenv("ARTICLE_AGENT_DISABLE_EVAL_CACHE", "0").strip().lower() in {"1", "true", "yes"}
    require_input_hash = os.getenv("ARTICLE_AGENT_REQUIRE_EVAL_INPUT_HASH", "0").strip().lower() in {"1", "true", "yes"}
    input_hash = hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest() if isinstance(payload, dict) else None
    force_synthesis = os.getenv("ARTICLE_AGENT_FORCE_EVAL_SYNTHESIS", "0").strip().lower() in {"1", "true", "yes"}
    if force_synthesis and isinstance(payload, dict):
        task = payload.get("task_description")
        task_text = task.get("task", "") if isinstance(task, dict) else str(task or "")
        field_defs = payload.get("field_definitions") if isinstance(payload.get("field_definitions"), dict) else {}
        if payload.get("module") == "outcomes" or field_defs.get("module") == "outcomes" or task_text.lower().startswith("synthesize"):
            disable_cache = True
    if output.exists() and not disable_cache:
        try:
            cached = json.loads(output.read_text(encoding="utf-8"))
        except (OSError, TypeError, ValueError):
            cached = None
        if require_input_hash and isinstance(cached, dict) and cached.get("_input_sha256") != input_hash:
            cached = None
        # A tagged outcome directory can contain artifacts from an earlier
        # run that used a different rows-per-request setting.  Reuse a cache
        # only when its row_audit/source_index set covers exactly this input;
        # otherwise the caller must issue a fresh request.  This prevents an
        # old part-0008 response from being mistaken for the current row-008
        # shard while still allowing a failed-only retry to skip valid parts.
        expected_indices = payload.get("field_definitions", {}).get("source_indices") if isinstance(payload, dict) else None
        if isinstance(expected_indices, list) and expected_indices:
            expected = {int(value) for value in expected_indices if value is not None}
            returned = []
            if isinstance(cached, dict):
                for key in ("row_audits", "audits", "records"):
                    values = cached.get(key)
                    if isinstance(values, list):
                        for item in values:
                            if isinstance(item, dict):
                                try:
                                    returned.append(int(item.get("source_index")))
                                except (TypeError, ValueError):
                                    pass
                        if returned:
                            break
            # Cache entries are accepted only when they describe exactly the
            # same source-index set as the current request.  A subset cache
            # can otherwise hide a partial response behind a successful
            # status and violate the lossless coverage contract.
            if expected == set(returned):
                return cached
        elif isinstance(cached, dict) and not cached.get("transport_failed"):
            return cached
    module_name = payload.get("field_definitions", {}).get("module") if isinstance(payload, dict) else None
    # Keep the transport-level role concise and English.  The user payload
    # still carries the requested Chinese medical-RCT role and field rules,
    # while a short system message avoids gateway timeouts on 5.6-sol.
    system_role = "You are an independent medical RCT extraction quality auditor. Return one JSON object only."
    # Never replace correctness-critical candidate/gold/source data with a
    # compact summary.  Request partitioning is handled by the outcome-row
    # evaluator instead of character clipping.
    wire_payload = payload
    messages = [
        {"role": "system", "content": system_role},
        {"role": "user", "content": json.dumps(wire_payload, ensure_ascii=False)},
    ]
    last_error = None
    for attempt in range(max(1, retries)):
        try:
            try:
                delay = max(0.0, float(request_delay_seconds))
            except (TypeError, ValueError):
                delay = 0.0
            if delay:
                time.sleep(delay)
            result = client.chat_json(messages)
            if isinstance(result, dict) and input_hash:
                result["_input_sha256"] = input_hash
            output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
            return result
        except RuntimeError as exc:
            last_error = exc
            time.sleep(2 * (attempt + 1))
    raise RuntimeError(f"LLM evaluation failed after {retries} attempts: {last_error}")


def _compact_api_prompts_enabled() -> bool:
    return os.getenv("ARTICLE_AGENT_COMPACT_API_PROMPTS", "0").strip().lower() in {"1", "true", "yes"}


def _clip_eval_value(value, max_chars: int = 700):
    """Compatibility helper that now preserves the complete value."""

    del max_chars
    return value


def _compact_eval_payload(payload: dict) -> dict:
    """Compatibility name for the old compact adapter; now lossless."""

    return payload


def _numeric_score(value: object) -> int | None:
    try:
        raw = float(value)
    except (TypeError, ValueError):
        return None
    # Some gateways/model aliases return a 0..1 quality proportion even
    # though the contract asks for 0..100.  Normalize that representation
    # before aggregation; integer 0 remains a legitimate zero score.
    if 0.0 < raw < 1.0:
        raw *= 100.0
    score = int(round(raw))
    return max(0, min(score, 100)) if 0 <= score <= 100 else None


def _finding_score(item: dict) -> int | None:
    """Derive a transparent 0–100 row score from an LLM's field findings.

    A few gateway/model responses copied the JSON template's ``module_score:
    0`` despite returning detailed findings.  In that case the raw score is
    not evidence of a completely unusable row.  This deterministic fallback
    uses only the returned finding statuses (not Gold values) and keeps the
    original score in ``module_score_original`` for audit.
    """

    findings = item.get("field_findings") if isinstance(item, dict) else None
    if not isinstance(findings, list) or not findings:
        return None
    weights = {
        "correct": 1.0,
        "acceptable": 0.8,
        "gold_ambiguous": 0.75,
        "missing": 0.3,
        "incorrect": 0.0,
        "partially_incorrect": 0.4,
    }
    values = []
    for finding in findings:
        if not isinstance(finding, dict):
            continue
        status = str(finding.get("status") or "").strip().lower()
        if status in weights:
            values.append(weights[status])
    if not values:
        return None
    return max(0, min(100, int(round(100.0 * sum(values) / len(values)))))


def _normalize_row_audit_score(item: dict) -> dict:
    """Normalize a row audit score while retaining the raw model value."""

    if not isinstance(item, dict):
        return item
    raw = item.get("module_score")
    numeric = _numeric_score(raw)
    try:
        raw_float = float(raw)
    except (TypeError, ValueError):
        raw_float = None
    if raw_float == 0.0 and item.get("field_findings"):
        derived = _finding_score(item)
        if derived is not None:
            item["module_score_original"] = raw
            item["module_score"] = derived
            item["score_source"] = "deterministic_mean_of_llm_field_findings"
            return item
    if numeric is not None:
        if raw_float is not None and 0.0 < raw_float < 1.0:
            item["module_score_original"] = raw
            item["score_source"] = "llm_fraction_normalized_to_percent"
        item["module_score"] = numeric
    return item


def _source_context_for_outcome(source_context: str, source: dict) -> str:
    """Return complete evidence unit for one outcome row, never a char clip."""

    evidence_quotes = source.get("evidence") if isinstance(source, dict) else []
    quotes = []
    if isinstance(evidence_quotes, list):
        for item in evidence_quotes:
            if isinstance(item, dict) and item.get("quote"):
                quotes.append(" ".join(str(item["quote"]).split()).lower())
    table_matches = list(re.finditer(r"<table\b.*?</table>", source_context, flags=re.I | re.S))
    for match in table_matches:
        table_text = match.group(0)
        normalized_table = " ".join(re.sub(r"<[^>]+>", " ", table_text).split()).lower()
        if quotes and any(quote in normalized_table for quote in quotes):
            return table_text
    normalized_context = " ".join(source_context.split()).lower()
    if quotes:
        for quote in quotes:
            index = normalized_context.find(quote)
            if index >= 0:
                # Preserve the complete paragraph containing the quoted
                # evidence.  This is source-unit routing, not truncation.
                for paragraph in re.split(r"\n\s*\n", source_context):
                    if quote in " ".join(paragraph.split()).lower():
                        return paragraph
    return source_context


def _lossless_table_evidence_contexts(extraction_dir: Path) -> dict[str, str]:
    """Build compact, lossless structural evidence units for table audits.

    The full Markdown/HTML tables remain in the article artifacts and in the
    extraction request inputs.  For evaluation, repeating a 40-KB table once
    per outcome row causes avoidable gateway timeouts.  The tablewise manifest
    contains the complete deterministic header/column map; each candidate
    record separately carries its verbatim source row.  Combining those two
    pieces gives the auditor all semantic evidence without duplicating the
    whole table payload.
    """

    path = extraction_dir / "raw_module_responses" / "outcomes.tablewise.manifest.json"
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, TypeError, ValueError):
        return {}
    tables = payload.get("tables") if isinstance(payload, dict) else None
    if not isinstance(tables, list):
        return {}
    contexts: dict[str, str] = {}
    for table in tables:
        if not isinstance(table, dict):
            continue
        table_id = str(table.get("table_id") or "").strip()
        if not table_id:
            continue
        lines = [
            f"TABLE_ID: {table_id}",
            f"TABLE_CAPTION: {table.get('caption') or 'NR'}",
            f"TABLE_CATEGORY: {table.get('table_category') or 'unknown'}",
            f"TABLE_SOURCE: {table.get('source') or 'NR'}",
            f"TARGET_SELECTION_REASON: {table.get('selection_reason') or 'NR'}",
            "TABLE_COLUMN_LABELS (complete deterministic labels):",
            json.dumps(table.get("column_labels") or [], ensure_ascii=False, separators=(",", ":")),
            "TABLE_ARM_REGISTRY (complete deterministic group labels/sample sizes):",
            json.dumps(table.get("arm_registry") or [], ensure_ascii=False, separators=(",", ":")),
            "TABLE_TIMEPOINT_LABELS (complete deterministic labels):",
            json.dumps(table.get("timepoint_labels") or [], ensure_ascii=False, separators=(",", ":")),
            "TABLE_STATISTIC_COLUMNS (complete deterministic statistic headers):",
            json.dumps(table.get("statistic_columns") or [], ensure_ascii=False, separators=(",", ":")),
            "TABLE_COLUMN_MAP (complete multilevel header paths, indexes, arms, timepoints, statistics, evidence):",
            json.dumps(table.get("column_map") or [], ensure_ascii=False, separators=(",", ":")),
            f"TABLE_SOURCE_ROW_COUNT: {table.get('row_count', 'NR')}",
            f"TABLE_SELECTED_ROW_COUNT: {table.get('selected_row_count', 'NR')}",
        ]
        contexts[table_id] = "\n".join(lines)
    return contexts


def evaluate_outcomes_by_row(
    client: OpenAICompatibleClient,
    candidate: dict,
    gold_rows: list[dict],
    source_context: str,
    extraction_dir: Path,
    *,
    tag: str = "",
    retries: int = 2,
    request_delay_seconds: float = 0.01,
    table_evidence_contexts: dict[str, str] | None = None,
) -> dict:
    """Audit every outcome source record without a summary-size shortcut.

    Each row receives a complete source record, complete routed Results
    context, and the complete Sheet3 reference.  The final module score is
    synthesized from all row audits; if the synthesis transport fails, the
    deterministic mean of the successful row audits is recorded as a
    provenance-safe fallback rather than silently scoring a prefix.
    """

    records = candidate.get("outcomes", []) if isinstance(candidate, dict) else []
    records = [item for item in records if isinstance(item, dict)] if isinstance(records, list) else []
    suffix = f".{tag}" if tag else ""
    parts_dir = extraction_dir / "raw_module_responses" / "evaluation_outcomes"
    parts_dir.mkdir(parents=True, exist_ok=True)
    part_manifest: list[dict] = []
    row_audits: list[dict] = []
    row_inputs: list[dict] = []

    try:
        rows_per_request = int(os.getenv("ARTICLE_AGENT_EVAL_OUTCOME_ROWS_PER_REQUEST", "1"))
    except ValueError:
        rows_per_request = 1
    rows_per_request = max(1, min(rows_per_request, 16))
    # An empty candidate is still audited once, so a missing outcome dataset
    # cannot accidentally receive a perfect score from an empty loop.
    indexed_records = list(enumerate(records)) or [(None, {})]
    batches = [
        indexed_records[offset:offset + rows_per_request]
        for offset in range(0, len(indexed_records), rows_per_request)
    ]
    for part_index, batch_items in enumerate(batches, start=1):
        source_indices = [item[0] for item in batch_items]
        source_records = [item[1] for item in batch_items]
        # Keep one complete evidence unit per distinct table/paragraph and
        # reference it from every source row.  Earlier payloads duplicated a
        # whole table once per row; that was semantically lossless but made
        # large evaluation requests needlessly exceed gateway limits.  The
        # unit itself is never clipped, and the per-row refs preserve exact
        # source_index/table_id/row_id coverage for auditability.
        source_context_refs = []
        source_context_units = []
        unit_by_key = {}
        for source_index, record in batch_items:
            source = record.get("source_outcome") if isinstance(record.get("source_outcome"), dict) else record
            table_id = source.get("table_id", "NR") if isinstance(source, dict) else "NR"
            row_id = source.get("row_id", "NR") if isinstance(source, dict) else "NR"
            source_table_id = str(table_id)
            context = (
                (table_evidence_contexts or {}).get(source_table_id)
                or _source_context_for_outcome(source_context, source if isinstance(source, dict) else {})
            )
            unit_key = (str(table_id), context)
            unit_id = unit_by_key.get(unit_key)
            if unit_id is None:
                unit_id = f"u{len(source_context_units) + 1:04d}"
                unit_by_key[unit_key] = unit_id
                source_context_units.append({
                    "unit_id": unit_id,
                    "table_id": table_id,
                    "row_ids": [],
                    "source_indices": [],
                    "context": context,
                })
            unit = source_context_units[-1] if source_context_units[-1]["unit_id"] == unit_id else next(
                item for item in source_context_units if item["unit_id"] == unit_id
            )
            unit["row_ids"].append(row_id)
            unit["source_indices"].append(source_index)
            source_context_refs.append({
                "source_index": source_index,
                "table_id": table_id,
                "row_id": row_id,
                "unit_id": unit_id,
            })
        payload = {
            "role_definition": "医学 RCT 结局抽取质量审计专家；只返回 JSON，不修改来源值。",
            "task_description": "逐一评价全部 source_index；依据完整来源记录、表头结构、行证据和 Gold 评分。允许证据唯一支持的推导，禁止猜测、借值或改写 source_outcome。必须覆盖全部索引；每条只返回0–100整数分和不超过30字的简短结论，不展开长篇逐字段解释。",
            "field_definitions": {
                "source_indices": source_indices,
                "candidate_records": "完整原始记录；source_outcome 优先于 normalized 字段",
                "human_gold_values": "完整 Sheet3，仅用于评价和冲突标记",
                "source_context_units": "用 unit_id 解析完整多级表头/来源结构",
            },
            "candidate_records": [
                {"source_index": source_index, "record": record}
                for source_index, record in zip(source_indices, source_records)
            ],
            # The list forms are canonical.  Do not duplicate the complete
            # source record for a one-row shard; duplicated aliases previously
            # doubled a 40-KB evidence table without adding information.
            "human_gold_values": [dict(row) for row in gold_rows if isinstance(row, dict)],
            "source_contexts": source_context_refs,
            "source_context_units": source_context_units,
            "source_contexts_rule": "按 unit_id 解析；context 完整不截断",
            "lossless_input": True,
            "json_template": {
                "row_audits": [{
                    "source_index": source_index,
                    "module_score": 0,
                    "module_verdict": "不超过30字的证据结论",
                } for source_index in source_indices],
            },
        }
        input_path = parts_dir / f"evaluation_input{suffix}.outcomes.part-{part_index:04d}.json"
        input_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        row_inputs.append(payload)
        output_path = parts_dir / f"llm_evaluation{suffix}.outcomes.part-{part_index:04d}.json"

        def cached_row_fallback_audits() -> list[dict] | None:
            """Load a complete prior row-fallback set without reissuing the
            known-failing multi-row request.

            A previous run may have split this exact block after a partial
            response.  The row files carry their own input hashes, so using
            them is safe only when every current one-row payload hashes
            identically; otherwise the normal API/fallback path is used.
            """

            if len(batch_items) <= 1:
                return None
            cached_items: list[dict] = []
            for source_index, record in batch_items:
                row_payload = dict(payload)
                row_payload["field_definitions"] = dict(payload.get("field_definitions") or {})
                row_payload["field_definitions"]["source_indices"] = [source_index]
                row_payload["candidate_records"] = [{"source_index": source_index, "record": record}]
                row_payload["source_contexts"] = [
                    ref for ref in source_context_refs if ref.get("source_index") == source_index
                ]
                referenced_units = {ref.get("unit_id") for ref in row_payload["source_contexts"]}
                row_payload["source_context_units"] = [
                    unit for unit in source_context_units if unit.get("unit_id") in referenced_units
                ]
                row_payload["json_template"] = {
                    "row_audits": [{
                        "source_index": source_index,
                        "module_score": 0,
                        "field_findings": [{
                            "field": "string",
                            "status": "correct|acceptable|incorrect|missing|gold_ambiguous",
                            "reason": "evidence-based",
                        }],
                    }]
                }
                row_payload["fallback_from_part"] = part_index
                row_output = parts_dir / (
                    f"llm_evaluation{suffix}.outcomes.part-{part_index:04d}.row-{int(source_index):04d}.json"
                )
                try:
                    cached = json.loads(row_output.read_text(encoding="utf-8"))
                except (OSError, TypeError, ValueError):
                    return None
                expected_hash = hashlib.sha256(
                    json.dumps(row_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
                ).hexdigest()
                if not isinstance(cached, dict) or cached.get("_input_sha256") != expected_hash:
                    return None
                returned = cached.get("row_audits") or cached.get("audits") or cached.get("records")
                item = None
                if isinstance(returned, list):
                    for candidate_item in returned:
                        if not isinstance(candidate_item, dict):
                            continue
                        try:
                            if int(candidate_item.get("source_index")) == int(source_index):
                                item = dict(candidate_item)
                                break
                        except (TypeError, ValueError):
                            continue
                if item is None and any(key in cached for key in ("module_score", "field_findings", "module_verdict")):
                    item = dict(cached)
                if item is None:
                    return None
                item["source_index"] = source_index
                source = record.get("source_outcome") if isinstance(record.get("source_outcome"), dict) else record
                item["table_id"] = source.get("table_id", "NR") if isinstance(source, dict) else "NR"
                item["row_id"] = source.get("row_id", "NR") if isinstance(source, dict) else "NR"
                item["fallback_from_part"] = part_index
                cached_items.append(_normalize_row_audit_score(item))
            return cached_items if len(cached_items) == len(batch_items) else None

        try:
            audit = None
            cached_fallback = cached_row_fallback_audits()
            if cached_fallback is not None:
                audit = {"row_audits": cached_fallback, "_cached_row_fallback": True}
            if audit is None:
                audit = call_judge(client, payload, output_path, retries=retries, request_delay_seconds=request_delay_seconds)
            if not isinstance(audit, dict):
                audit = {"module_score": 0, "module_verdict": "invalid non-object judge response", "field_findings": []}
            audit = dict(audit)
            # Multi-row shards should return a row_audits/audits list.  If a
            # gateway returns only a shard-level score, retain that score as a
            # conservative annotation for every source row rather than
            # silently dropping the covered indices; the raw response remains
            # available for review.
            returned_audits = audit.get("row_audits") or audit.get("audits") or audit.get("records")
            if not isinstance(returned_audits, list):
                returned_audits = []
            by_index = {}
            for item in returned_audits:
                if not isinstance(item, dict):
                    continue
                try:
                    idx = int(item.get("source_index"))
                except (TypeError, ValueError):
                    continue
                by_index[idx] = dict(item)
            expected_index_set = set(source_indices)
            returned_index_set = set(by_index)
            if returned_index_set != expected_index_set:
                # A few compatible gateways return a single-row audit as a
                # compact object (module_score/module_verdict/field_findings)
                # instead of wrapping it in ``row_audits``.  The request has
                # exactly one declared source_index, so attaching that index
                # is deterministic and auditable; multi-row shards never use
                # this compatibility path and must return every index.
                if (
                    len(batch_items) == 1
                    and not returned_index_set
                    and isinstance(audit, dict)
                    and any(key in audit for key in ("module_score", "field_findings", "module_verdict"))
                ):
                    by_index[source_indices[0]] = dict(audit)
                    returned_index_set = expected_index_set
                else:
                    missing = sorted(expected_index_set - returned_index_set)
                    extra = sorted(returned_index_set - expected_index_set)
                    raise ValueError(
                        "outcome evaluation response source_index set mismatch: "
                        f"missing={missing}, extra={extra}"
                    )
            for source_index, record in batch_items:
                item = by_index.get(source_index)
                if item is None:
                    item = {
                        "module_score": audit.get("module_score", 0),
                        "module_verdict": audit.get("module_verdict", "shard-level audit; source row retained"),
                        "field_findings": audit.get("field_findings", []),
                        "strengths": audit.get("strengths", []),
                        "weaknesses": audit.get("weaknesses", []),
                        "gold_quality_notes": audit.get("gold_quality_notes", []),
                        "shard_level_fallback": len(batch_items) > 1,
                    }
                item["source_index"] = source_index
                source = record.get("source_outcome") if isinstance(record.get("source_outcome"), dict) else record
                item["table_id"] = source.get("table_id", "NR") if isinstance(source, dict) else "NR"
                item["row_id"] = source.get("row_id", "NR") if isinstance(source, dict) else "NR"
                item = _normalize_row_audit_score(item)
                row_audits.append(item)
            part_manifest.append({
                "part": part_index,
                "source_indices": source_indices,
                "row_count": len(source_indices),
                "table_ids": [item.get("table_id") for item in row_audits if item.get("source_index") in source_indices],
                "row_ids": [item.get("row_id") for item in row_audits if item.get("source_index") in source_indices],
                "input_sha256": hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest(),
                "lossless": True,
                "status": "success",
                "score": _numeric_score(audit.get("module_score")),
            })
        except Exception as exc:
            # A multi-row shard can fail because its response is just over a
            # gateway budget even though each complete row is independently
            # evaluable.  Fall back to one-row requests, preserving the exact
            # candidate/evidence payload and recording every attempt.  A
            # one-row failure remains an explicit failed item for retry.
            fallback_enabled = os.getenv("ARTICLE_AGENT_EVAL_ROW_FALLBACK", "1").strip().lower() in {"1", "true", "yes"}
            fallback_audits = []
            fallback_failures = []
            if fallback_enabled and len(batch_items) > 1:
                for source_index, record in batch_items:
                    row_payload = dict(payload)
                    row_payload["field_definitions"] = dict(payload.get("field_definitions") or {})
                    row_payload["field_definitions"]["source_indices"] = [source_index]
                    row_payload["candidate_records"] = [{"source_index": source_index, "record": record}]
                    row_payload["source_contexts"] = [
                        ref for ref in source_context_refs if ref.get("source_index") == source_index
                    ]
                    referenced_units = {ref.get("unit_id") for ref in row_payload["source_contexts"]}
                    row_payload["source_context_units"] = [
                        unit for unit in source_context_units if unit.get("unit_id") in referenced_units
                    ]
                    row_payload["json_template"] = {
                        "row_audits": [{
                            "source_index": source_index,
                            "module_score": 0,
                            "field_findings": [{"field": "string", "status": "correct|acceptable|incorrect|missing|gold_ambiguous", "reason": "evidence-based"}],
                        }]
                    }
                    row_payload["fallback_from_part"] = part_index
                    row_output = parts_dir / f"llm_evaluation{suffix}.outcomes.part-{part_index:04d}.row-{int(source_index):04d}.json"
                    try:
                        row_result = call_judge(
                            client,
                            row_payload,
                            row_output,
                            retries=retries,
                            request_delay_seconds=request_delay_seconds,
                        )
                        returned = row_result.get("row_audits") or row_result.get("audits") or row_result.get("records") if isinstance(row_result, dict) else None
                        item = None
                        if isinstance(returned, list):
                            for candidate_item in returned:
                                if not isinstance(candidate_item, dict):
                                    continue
                                try:
                                    if int(candidate_item.get("source_index")) == int(source_index):
                                        item = dict(candidate_item)
                                        break
                                except (TypeError, ValueError):
                                    continue
                        if item is None:
                            item = dict(row_result) if isinstance(row_result, dict) else {}
                        item["source_index"] = source_index
                        source = record.get("source_outcome") if isinstance(record.get("source_outcome"), dict) else record
                        item["table_id"] = source.get("table_id", "NR") if isinstance(source, dict) else "NR"
                        item["row_id"] = source.get("row_id", "NR") if isinstance(source, dict) else "NR"
                        item["fallback_from_part"] = part_index
                        item = _normalize_row_audit_score(item)
                        fallback_audits.append(item)
                    except Exception as fallback_exc:
                        fallback_failures.append({"source_index": source_index, "error": str(fallback_exc)})
            row_audits.extend(fallback_audits)
            source = source_records[0].get("source_outcome") if source_records and isinstance(source_records[0].get("source_outcome"), dict) else (source_records[0] if source_records else {})
            fallback_scores = [_numeric_score(item.get("module_score")) for item in fallback_audits]
            fallback_scores = [score for score in fallback_scores if score is not None]
            all_fallback_succeeded = bool(fallback_audits) and not fallback_failures and len(fallback_audits) == len(batch_items)
            part_manifest.append({
                "part": part_index,
                "source_indices": source_indices,
                "row_count": len(source_indices),
                "table_id": source.get("table_id", "NR") if isinstance(source, dict) else "NR",
                "row_id": source.get("row_id", "NR") if isinstance(source, dict) else "NR",
                "input_sha256": hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest(),
                "lossless": True,
                "status": "success" if all_fallback_succeeded else "partial" if fallback_audits else "failed",
                "fallback_mode": "row" if len(batch_items) > 1 else None,
                "covered_source_indices": [item.get("source_index") for item in fallback_audits],
                "failed_source_indices": [item.get("source_index") for item in fallback_failures] or ([] if fallback_audits else source_indices),
                "score": int(round(sum(fallback_scores) / len(fallback_scores))) if fallback_scores else None,
                "error": str(exc),
                "fallback_errors": fallback_failures,
            })

    final_payload = {
        "role_definition": "你是一名医学 RCT 抽取质量综合审计专家。",
        "task_description": (
            "综合当前文章所有逐行 outcome audit。必须覆盖每个 source_index，不得只依据前几条。"
            "以论文证据为最高依据，区分候选错误、Gold 错误和证据不足；source_outcome 原始值不可被规范化字段替换。"
            "module_score 为所有已审计记录和未匹配 Gold 行的整体结局质量评分；返回 JSON。"
        ),
        "module": "outcomes",
        "candidate": {"outcomes": records, "record_count": len(records)},
        "human_gold_values": [dict(row) for row in gold_rows if isinstance(row, dict)],
        "source_context": source_context,
        "row_audits": row_audits,
        "row_audit_manifest": part_manifest,
        "lossless_input": True,
        "json_template": {
            "module_score": 0,
            "module_verdict": "string",
            "field_findings": [],
            "strengths": [],
            "weaknesses": [],
            "gold_quality_notes": [],
        },
    }
    final_input = extraction_dir / f"evaluation_input{suffix}.outcomes.json"
    final_input.write_text(json.dumps(final_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    final_output = extraction_dir / f"llm_evaluation{suffix}.outcomes.json"
    valid_scores = [_numeric_score(item.get("module_score")) for item in row_audits]
    valid_scores = [score for score in valid_scores if score is not None]
    skip_synthesis = os.getenv("ARTICLE_AGENT_SKIP_EVAL_SYNTHESIS", "0").strip().lower() in {
        "1", "true", "yes"
    }
    if skip_synthesis:
        # The row audits are already the correctness-critical evaluation.  A
        # second whole-article request can exceed gateway body/time limits;
        # when explicitly disabled, keep a transparent deterministic
        # synthesis artifact instead of issuing an unbounded request.
        synthesis = {
            "module_score": int(round(sum(valid_scores) / len(valid_scores))) if valid_scores else 0,
            "module_verdict": "deterministic mean over complete row audits; synthesis request explicitly skipped",
            "field_findings": [],
            "strengths": [],
            "weaknesses": [],
            "gold_quality_notes": [],
            "synthesis_fallback": True,
            "synthesis_skip_reason": "ARTICLE_AGENT_SKIP_EVAL_SYNTHESIS=1",
        }
    else:
        try:
            synthesis = call_judge(client, final_payload, final_output, retries=retries, request_delay_seconds=request_delay_seconds)
        except Exception as exc:
            synthesis = {
                "module_score": int(round(sum(valid_scores) / len(valid_scores))) if valid_scores else 0,
                "module_verdict": "row-level synthesis transport failed; deterministic mean fallback",
                "field_findings": [],
                "strengths": [],
                "weaknesses": [str(exc)],
                "gold_quality_notes": [],
                "synthesis_fallback": True,
            }
    if not isinstance(synthesis, dict):
        synthesis = {}
    if _numeric_score(synthesis.get("module_score")) is None:
        synthesis["module_score"] = int(round(sum(valid_scores) / len(valid_scores))) if valid_scores else 0
        synthesis["synthesis_fallback"] = True
    # The authoritative outcome score is the deterministic aggregate of every
    # expected source_index.  A synthesis model may provide a useful narrative
    # score, but it must not silently ignore failed/missing shards or choose a
    # prefix.  Missing row audits contribute zero and the model's value is
    # retained separately for comparison.
    score_by_index = {}
    for item in row_audits:
        if not isinstance(item, dict):
            continue
        try:
            index = int(item.get("source_index"))
        except (TypeError, ValueError):
            continue
        value = _numeric_score(item.get("module_score"))
        if value is not None:
            score_by_index[index] = value
    expected_index_set = set(range(len(records)))
    aggregate_scores = [score_by_index.get(index, 0) for index in range(len(records))]
    if aggregate_scores:
        model_score = _numeric_score(synthesis.get("module_score"))
        synthesis["model_synthesis_score"] = model_score
        synthesis["module_score"] = int(round(sum(aggregate_scores) / len(aggregate_scores)))
        synthesis["score_aggregation"] = {
            "method": "deterministic_mean_over_all_source_indices",
            "expected_source_index_count": len(expected_index_set),
            "covered_source_index_count": len(expected_index_set.intersection(score_by_index)),
            "missing_source_indices": sorted(expected_index_set - set(score_by_index)),
            "failed_rows_contribute_zero": True,
            "lossless": True,
        }
    synthesis["row_audits"] = row_audits
    synthesis["row_audit_manifest"] = part_manifest
    synthesis["lossless"] = True
    synthesis.setdefault("run_id", os.getenv("ARTICLE_AGENT_RUN_ID") or extraction_dir.parent.name)
    synthesis.setdefault("evaluation_model", getattr(client, "model", ""))
    final_output.write_text(json.dumps(synthesis, ensure_ascii=False, indent=2), encoding="utf-8")
    (parts_dir / f"outcomes{suffix}.manifest.json").write_text(
        json.dumps({
            "strategy": "lossless_one_source_row_per_judge_then_full_synthesis",
            "lossless": True,
            "source_record_count": len(records),
            "row_audit_count": len(row_audits),
            "rows_per_request": rows_per_request,
            "successful_row_audits": len(valid_scores),
            "parts": part_manifest,
        }, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return synthesis


def evaluate(
    extraction_dir: Path,
    tag: str = "",
    use_postprocessed: bool = False,
    request_delay_seconds: float | None = None,
) -> dict:
    if request_delay_seconds is None:
        try:
            request_delay_seconds = float(os.getenv("ARTICLE_AGENT_OUTCOME_REQUEST_DELAY_SECONDS", "0.01"))
        except ValueError:
            request_delay_seconds = 0.01
    request_delay_seconds = max(0.0, min(float(request_delay_seconds), 60.0))
    extraction = json.loads((extraction_dir / "extraction.json").read_text(encoding="utf-8"))
    contexts = json.loads((extraction_dir / "routed_context.json").read_text(encoding="utf-8"))
    try:
        article_manifest = json.loads((extraction_dir / "manifest.json").read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        article_manifest = {}
    article_id = extraction.get("article_id") or extraction_dir.name
    run_id = str(article_manifest.get("run_id") or os.getenv("ARTICLE_AGENT_RUN_ID") or extraction_dir.parent.name)
    gold1 = sheet1_gold(article_id)
    module_gold_fields = {
        "metadata": {"title", "publication_year", "journal", "first_author", "country", "intervention", "control"},
        "acupuncture": {"control_type_transformed", "treatment_frequency_raw", "treatment_frequency_value", "treatment_frequency_unit", "treatment_duration_raw", "treatment_duration_value", "treatment_duration_unit", "total_sessions", "deqi"},
        "risk_of_bias": {"random_sequence_method", "random_sequence_class", "allocation_concealment", "allocation_concealment_class", "participant_blinding", "outcome_assessor_blinding", "primary_analysis", "missing_data_method"},
        "consort_flow": {"randomized_sample_intervention_raw", "randomized_sample_control_raw", "total_randomized"},
    }
    module_sources = {
        "metadata": contexts["metadata"],
        "acupuncture": evidence_windows(contexts["acupuncture"], [
            "acupuncture", "sham", "needle", "treatment", "session", "de qi", "deqi",
        ], radius=350, limit=1400),
        "risk_of_bias": evidence_windows(contexts["risk_of_bias"], [
            "random", "allocation", "conceal", "blind", "intention-to-treat", "missing",
        ], radius=350, limit=1400),
        "outcomes": evidence_windows(contexts["outcomes"], [
            "primary outcome", "table", "p=", "p <", "confidence interval", "baseline",
        ], radius=300, limit=900),
        "consort_flow": extraction.get("consort_flow", {}).get("evidence", []) if extraction.get("consort_flow") else [],
    }
    module_codebooks = {
        "acupuncture": {
            "control_type_transformed": "1=penetrating sham, 2=non-penetrating sham, 3=non-needle sham, 4=high-intensity non-sham, 5=usual care, 6=low-intensity non-sham",
            "deqi": "1=yes, 2=no, 3=not reported, 4=not applicable",
            "treatment_frequency_unit": "1=day, 2=week, 3=hour",
            "treatment_duration_unit": "1=day, 2=week",
        },
        "risk_of_bias": {
            "random_sequence_class": "1=random number table, 2=computer random, 3=lottery, 4=dice, 5=coin, 6=shuffle/envelope, 7=minimization, 8=method not reported, 9=other",
            "allocation_concealment_class": "1=central telephone/web, 2=opaque sealed envelope, 3=sealed envelope, 4=opaque envelope, 5=not reported, 6=other",
            "participant_blinding/outcome_assessor_blinding": "1=yes, 2=no, 3=not reported",
            "primary_analysis": "1=ITT/mITT, 2=available case, 3=per protocol, 4=no explicit primary analysis",
            "missing_data_method": "1=complete case, 2=all available, 3=mean, 4=LOCF, 5=regression, 6=multiple imputation, 7=maximum likelihood, 8=weighting, 9=combination, 10=mixed effect, 11=other, 12=no missing, 13=not reported",
        },
    }
    outcome_candidate = extraction["outcomes"]
    postprocessed_used = False
    if use_postprocessed:
        outcome_candidate, postprocessed_used = postprocessed_outcome_candidate(
            extraction_dir, extraction["outcomes"]
        )
    outcome_table_evidence_contexts = _lossless_table_evidence_contexts(extraction_dir)
    module_candidates = {
        "metadata": extraction["metadata"],
        "acupuncture": extraction["acupuncture"],
        "risk_of_bias": extraction["risk_of_bias"],
        "outcomes": outcome_candidate,
        "consort_flow": extraction.get("consort_flow"),
    }
    try:
        eval_timeout = int(os.getenv("ARTICLE_AGENT_EVAL_TIMEOUT", "30"))
    except ValueError:
        eval_timeout = 30
    try:
        eval_retries = int(os.getenv("ARTICLE_AGENT_EVAL_RETRIES", "1"))
    except ValueError:
        eval_retries = 1
    # Evaluation payloads are compact and the gateway health checks complete
    # within a few seconds.  Allow a shorter lower bound so one unavailable
    # base URL cannot stall an article for several minutes during failover.
    eval_timeout = max(10, min(eval_timeout, 300))
    eval_retries = max(1, min(eval_retries, 5))
    eval_model = os.getenv("ARTICLE_AGENT_EVAL_MODEL") or os.getenv("ARTICLE_AGENT_MODEL")
    client = OpenAICompatibleClient(model=eval_model, timeout=eval_timeout)
    module_results = {}
    common = {
        "authority_order": [
            "Quoted article/figure evidence is the ultimate authority.",
            "Human Excel gold is a reference and may contain coding or year/sample-size inconsistencies.",
            "The candidate extraction must not receive credit merely for being non-null.",
        ],
        "field_status": ["correct", "acceptable", "incorrect", "missing", "gold_ambiguous"],
        "required_json": {
            "module_score": 0,
            "module_verdict": "string",
            "field_findings": [{
                "field": "string", "status": "correct|acceptable|incorrect|missing|gold_ambiguous",
                "severity": "critical|major|minor", "reason": "brief string",
            }],
            "strengths": ["string"],
            "weaknesses": ["string"],
            "gold_quality_notes": ["string"],
        },
    }
    for module in ("metadata", "acupuncture", "risk_of_bias", "outcomes", "consort_flow"):
        gold = sheet3_gold(article_id) if module == "outcomes" else {
            key: value for key, value in gold1.items() if key in module_gold_fields.get(module, set())
        }
        compact_gold = compact_gold_rows_for_evaluation(gold) if isinstance(gold, list) else {
            key: (value.get("value") if isinstance(value, dict) and "value" in value else value)
            for key, value in gold.items()
        }
        candidate = module_candidates[module]
        if module == "outcomes":
            # Outcomes are correctness-critical and are audited one source row
            # at a time.  The helper writes full row inputs and a complete
            # synthesis artifact; no summary prefix is sent to the judge.
            module_results[module] = evaluate_outcomes_by_row(
                client,
                candidate,
                gold if isinstance(gold, list) else [],
                module_sources[module],
                extraction_dir,
                tag=tag,
                retries=eval_retries,
                request_delay_seconds=request_delay_seconds,
                table_evidence_contexts=outcome_table_evidence_contexts,
            )
            continue
        elif module == "consort_flow":
            # Flow evidence is also sent in full; the figure-specific source
            # list stays intact for an independent VLM audit.
            candidate = candidate if isinstance(candidate, dict) else {}
        elif isinstance(candidate, dict):
            candidate = dict(candidate)
        source_context = module_sources[module]
        if module == "consort_flow" and isinstance(source_context, list):
            source_context = [dict(item) for item in source_context if isinstance(item, dict)]
        payload = {
            "role_definition": "医学 RCT 论文信息提取专家兼独立质量审计者",
            "task_description": {
                "task": f"独立评价文章 {article_id} 的 {module} 模块抽取质量",
                "authority_order": common["authority_order"],
                "scoring_rule": "module_score 必须为0至100的整数",
                "response_limit": "不限制字段发现条数；必须覆盖全部输入字段并返回完整 JSON",
            },
            "field_definitions": {
                "module": module,
                "fields_to_evaluate": sorted(module_gold_fields.get(module, set())),
                "codebooks": module_codebooks.get(module, {}),
                "field_status": common["field_status"],
                "candidate": candidate,
                "candidate_source": (
                    "outcomes.postprocessed.json (source_outcome + postprocess annotations)"
                    if module == "outcomes" and postprocessed_used
                    else "extraction.json"
                ),
                "human_gold_values": compact_gold,
                "source_context": source_context,
            },
            "json_template": common["required_json"],
        }
        suffix = f".{tag}" if tag else ""
        (extraction_dir / f"evaluation_input{suffix}.{module}.json").write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        module_output = extraction_dir / f"llm_evaluation{suffix}.{module}.json"
        try:
            module_results[module] = call_judge(
                client,
                payload,
                module_output,
                retries=eval_retries,
                request_delay_seconds=request_delay_seconds,
            )
        except Exception as exc:
            # Keep the article evaluation complete even when one module's
            # transport is unavailable.  This is an explicit failed audit,
            # not a score inferred from Gold; a later cache-enabled retry
            # detects ``transport_failed`` and reissues only this module.
            module_results[module] = {
                "module_score": 0,
                "module_verdict": "LLM transport failed; module requires retry",
                "field_findings": [],
                "strengths": [],
                "weaknesses": [str(exc)],
                "gold_quality_notes": [],
                "transport_failed": True,
            }
            module_output.write_text(json.dumps(module_results[module], ensure_ascii=False, indent=2), encoding="utf-8")

    # Keep every module finding in the synthesis payload.  In particular, the
    # full outcome row-audit list is required to calculate a batch-faithful
    # overall score; no prefix or character budget is applied here.
    compact_audits = module_results
    synthesis = {
        "role_definition": "Medical RCT extraction quality synthesis auditor",
        "task_description": {
            "task": "Synthesize five independent module audits into a final extraction-quality score",
            "scoring_rules": [
                "All module_scores and overall_score must be integers from 0 to 100",
                "Recalculate scores when module findings justify it",
                "Weight evidence errors in outcomes and risk of bias more heavily",
                "Distinguish candidate extraction errors from human gold errors",
            ],
        },
        "field_definitions": {"module_audits": compact_audits},
        "json_template": {
            "overall_score": 0,
            "module_scores": {name: 0 for name in module_results},
            "verdict": "string",
            "critical_errors": ["string"],
            "gold_quality_notes": ["string"],
            "next_actions": ["string"],
        },
    }
    suffix = f".{tag}" if tag else ""
    synthesis_output = extraction_dir / f"llm_evaluation{suffix}.json"
    # A full article-level synthesis payload can be substantially larger than
    # an individual module/row request.  When the caller explicitly enables
    # the lossless no-truncation mode, do not issue a potentially unbounded
    # synthesis request: the independent module audits are already complete,
    # so aggregate them deterministically and retain a transparent marker.
    skip_synthesis = os.getenv("ARTICLE_AGENT_SKIP_EVAL_SYNTHESIS", "0").strip().lower() in {
        "1", "true", "yes", "on"
    }
    if skip_synthesis:
        weights = {"metadata": 1.0, "acupuncture": 1.0, "risk_of_bias": 1.5, "outcomes": 2.0, "consort_flow": 1.0}
        fallback_scores = {
            name: (_numeric_score(audit.get("module_score")) if isinstance(audit, dict) else None)
            for name, audit in module_results.items()
        }
        valid = [(score, weights.get(name, 1.0)) for name, score in fallback_scores.items() if score is not None]
        result = {
            "overall_score": int(round(sum(score * weight for score, weight in valid) / sum(weight for _, weight in valid))) if valid else 0,
            "module_scores": {name: score for name, score in fallback_scores.items() if score is not None},
            "verdict": "deterministic weighted module-score aggregation; synthesis request explicitly skipped",
            "critical_errors": [],
            "gold_quality_notes": [],
            "next_actions": ["Inspect per-module and per-row audit artifacts."],
            "synthesis_fallback": True,
            "synthesis_skip_reason": "ARTICLE_AGENT_SKIP_EVAL_SYNTHESIS=1",
        }
        synthesis_output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    else:
        try:
            result = call_judge(
                client,
                synthesis,
                synthesis_output,
                retries=eval_retries,
                request_delay_seconds=request_delay_seconds,
            )
        except Exception as exc:
            # The module audits are independent and complete even when a single
            # giant synthesis request exceeds a gateway body limit.  Compute a
            # deterministic weighted fallback from those audits and retain the
            # transport error in the artifact instead of failing the article.
            weights = {"metadata": 1.0, "acupuncture": 1.0, "risk_of_bias": 1.5, "outcomes": 2.0, "consort_flow": 1.0}
            fallback_scores = {
                name: (_numeric_score(audit.get("module_score")) if isinstance(audit, dict) else None)
                for name, audit in module_results.items()
            }
            valid = [(score, weights.get(name, 1.0)) for name, score in fallback_scores.items() if score is not None]
            result = {
                "overall_score": int(round(sum(score * weight for score, weight in valid) / sum(weight for _, weight in valid))) if valid else 0,
                "module_scores": {name: score for name, score in fallback_scores.items() if score is not None},
                "verdict": "synthesis transport failed; weighted module-score fallback",
                "critical_errors": [str(exc)],
                "gold_quality_notes": [],
                "next_actions": ["Inspect per-module and per-row audit artifacts."],
                "synthesis_fallback": True,
            }
            synthesis_output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    # Guard against a truncated/partial synthesis response.  The five
    # module-level audits were already obtained independently, so filling a
    # missing module score from that audit is provenance-safe and does not use
    # the gold standard to alter extraction values.  Recalculate the overall
    # score only when the synthesis omitted at least one module.
    if isinstance(result, dict):
        raw_scores = result.get("module_scores")
        scores = dict(raw_scores) if isinstance(raw_scores, dict) else {}
        missing_modules = []
        for name, audit in module_results.items():
            value = scores.get(name)
            try:
                numeric = int(round(float(value)))
            except (TypeError, ValueError):
                numeric = None
            if numeric is None or not 0 <= numeric <= 100:
                try:
                    numeric = int(round(float((audit or {}).get("module_score"))))
                except (TypeError, ValueError, AttributeError):
                    numeric = None
            if numeric is None:
                missing_modules.append(name)
            else:
                if name not in scores:
                    missing_modules.append(name)
                scores[name] = max(0, min(numeric, 100))
        if missing_modules:
            weights = {
                "metadata": 1.0,
                "acupuncture": 1.0,
                "risk_of_bias": 1.5,
                "outcomes": 2.0,
                "consort_flow": 1.0,
            }
            weighted = [
                (score, weights.get(name, 1.0))
                for name, score in scores.items()
                if name in module_results and isinstance(score, int)
            ]
            if weighted:
                result["overall_score"] = int(round(sum(score * weight for score, weight in weighted) / sum(weight for _, weight in weighted)))
            result["module_scores"] = scores
            result["synthesis_fallback"] = {
                "reason": "synthesis_response_omitted_module_scores",
                "filled_modules": missing_modules,
                "weights": weights,
            }
            # Keep the audit artifact on disk synchronized with the returned
            # result so reports never mix a partial file with a complete
            # in-memory score.
            synthesis_output.write_text(
                json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
            )
    if isinstance(result, dict):
        result.setdefault("run_id", run_id)
        result.setdefault("evaluation_model", eval_model)
        result.setdefault("lossless", True)
        # The result was written once inside the transport/fallback branch.
        # Rewrite after provenance fields are attached so the on-disk artifact
        # cannot lose the run/model identity used by the batch report.
        synthesis_output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


if __name__ == "__main__":
    target = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else ROOT / "outputs/mineru_method_retest_eval_v2/2015-01"
    tag = sys.argv[2] if len(sys.argv) > 2 else ""
    use_postprocessed = "--use-postprocessed" in sys.argv[3:]
    result = evaluate(target, tag=tag, use_postprocessed=use_postprocessed)
    print(json.dumps({
        "overall_score": result.get("overall_score"),
        "module_scores": result.get("module_scores"),
        "verdict": result.get("verdict"),
        "output": str(target / f"llm_evaluation{'.' + tag if tag else ''}.json"),
    }, ensure_ascii=True, indent=2))
