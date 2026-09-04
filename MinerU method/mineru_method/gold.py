from __future__ import annotations

from pathlib import Path


def sheet3_gold(project_root: Path, article_id: str) -> list[dict]:
    """Load Sheet3 rows for post-extraction comparison only.

    The workbook is deliberately read after the source outcomes have already
    been extracted.  The returned rows are a comparison reference and are
    never used to construct the table-wise extraction prompt.
    """
    from openpyxl import load_workbook

    workbook_path = project_root / "Datas" / "label" / "2015-6篇-0813.xlsx"
    if not workbook_path.exists():
        return []
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet = workbook["Sheet3"]
        headers = {
            column: str(sheet.cell(5, column).value or f"column_{column}")
            for column in range(1, sheet.max_column + 1)
        }
        rows: list[dict] = []
        for row_index in range(1, sheet.max_row + 1):
            raw_id = str(sheet.cell(row_index, 1).value or "").strip()
            if raw_id != article_id and not raw_id.startswith(f"{article_id}-"):
                continue
            row = {
                key: sheet.cell(row_index, column).value
                for column, key in headers.items()
                if sheet.cell(row_index, column).value not in (None, "")
            }
            row["gold_row_id"] = f"{raw_id or article_id}-excel-row-{row_index:03d}"
            row["excel_row"] = row_index
            rows.append(row)
        return rows
    finally:
        workbook.close()
